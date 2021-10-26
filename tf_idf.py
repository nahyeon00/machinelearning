# from openpyxl import load_workbook
# from konlpy.tag import Komoran
# from sklearn.feature_extraction.text import TfidfVectorizer
# import pandas as pd
# import numpy as np
#
# df = load_workbook('데이터의 복사본.xlsx')  # 엑셀파일 열기
# data = df.active  # 시트 활성화(시트 하나뿐이기때문에 첫번째 시트 선택됨)
#
# komoran = Komoran()
# doc = list()
# results = []
#
# col = data['D']
# for cell in col:
#     if cell.value is None:
#         cell.value = "0"
#     doc.append(komoran.pos(cell.value))
#
#
#
# # NNG: 일반명사, NNP: 고유명사, NNB: 의존명사, NP: 대명사, NR: 수사
# # VV: 동사, VA: 형용사, MM: 관형사, SL: 외국어
# pos = ["NNG", "NNP", "NNB", "NP", "NR", "VV", "VA", "MM", "SL"]
#
# for i, sent in enumerate(doc):
#     words = [token[0] for token in sent if token[1] in pos]
#     results.append(" ".join(words))
#     if i >= 1:
#         data.cell(row=i+1, column=5).value = results[i]
#
# print(results)
#
# vectorizer = TfidfVectorizer()
# tfidfv = TfidfVectorizer().fit(results)  # tf-idf 객체선언 후 단어 학습시킴
# v = tfidfv.vocabulary_
# feature_names = tfidfv.get_feature_names()

# print(sorted(tfidfv.transform(results).toarray()[1], reverse=True))  # 코퍼스로부터 각 단어의 빈도수 기록,
#                                             # 단어 없이 빈도수만 출력 어떤 빈도수인지 알 수 없다.
# a = tfidfv.transform(results).toarray()
# data = {'tf-idf': a, 'vocabulary': v}
# df = pd.DataFrame(data)
# print(df)
# # feature_names = np.array(vectorizer.get_features_names())
# print(feature_names[a[:10]])


# vectorizer = TfidfVectorizer()
# X = vectorizer.fit_transform(results)
# # feature_array = vectorizer.get_feature_names()
# ziptwo = sorted(list(zip(vectorizer.get_feature_names(),
#                          X.sum(0).getA1())), key=lambda x:x[1], reverse=True)
# print(ziptwo[:5])




# a = tfidfv.transform(results)
# max_value = a.max(axis=1).toarray()
# sorted_by_tfidf = max_value.argsort()
# print(sorted_by_tfidf)


# print(tfidfv.vocabulary_)  # 각 단어의 인덱스가 어떻게 부여되었는지 확인
#                                    # 단어의 인덱스를 붙여주고 인덱스가 0부터 시작


# # sorted(tfidfv.vocabulary_items()) # 단어 사전 정렬
# last = tfidfv.transform(results)
# max_value = last.max(axis=0).toarray().ravel()
# sorted_by_tifidf = max_value.argsort()
# feature_names = np.array(tfidfv.get_feature_names())
# print(feature_names[sorted_by_tifidf[:10]])


# # 엑셀파일 저장
# df.save('데이터의 복사본.xlsx')
# df.close()
from collections import defaultdict
from openpyxl import load_workbook
from konlpy.tag import Komoran  # 품사 태깅 클래스
# 사이킥런의 feature_extraction.text 서브패키지 사용
# TfidfVectorizer: tf-idf 방식으로 단어의 가중치를 조정한 bow 인코딩 벡터를 만듬
from sklearn.feature_extraction.text import TfidfVectorizer

# 파일 이름이 데이터의 복사본.xlsx인 파일을 불러옴
df = load_workbook('데이터의 복사본.xlsx')
data = df.active  # 시트 활성화

komoran = Komoran()  # komoran
tfidf_vectorizer = TfidfVectorizer()  # TF-IDF 객체선언

doc = list()
results = []

col = data['D']
for cell in col:
    if cell.value is None:
        cell.value = "0"
    doc.append(komoran.pos(cell.value))  # pos(test): 텍스트에 품사 정보 부착해서 반환

# print(doc)  # [[('논문', 'NNP'), ('제목', 'NNG'),...,]]

pos = ["NNG", "NNP", "NNB", "NP", "NR", "VV", "VA", "MM", "SL"]
minus = []
for i, sent in enumerate(doc):
    words = [token[0] for token in sent if token[1] in pos]
    minus_set = set(words)
    minus.append(" ".join(minus_set))
    results.append(" ".join(words))

# TF-IDF
vectorizer = TfidfVectorizer()
sp_matrix = vectorizer.fit_transform(results)

word2id = defaultdict(lambda: 0)
for idx, feature in enumerate(vectorizer.get_feature_names()):
    word2id[feature] = idx

for i, sent in enumerate(minus):
    fin_words = ([(token, sp_matrix[i, word2id[token]]) for token in sent.split()])
    fin_words.sort(key=lambda x: x[1], reverse=True)
    fin_words = ([x[0] for x in fin_words])
    if i >= 1:
        data.cell(row=i+1, column=5).value = " ".join(fin_words)

# 엑셀 저장
df.save('데이터의 복사본.xlsx')
df.close()